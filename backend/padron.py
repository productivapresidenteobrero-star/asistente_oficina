import logging
import io
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from backend.database import get_db

logger = logging.getLogger("padron")
router = APIRouter(prefix="/api/padron", tags=["padron"])

# ── Schemas ──────────────────────────────────────────────────────────────────

class ConsejoCreate(BaseModel):
    nombre: str
    rif: Optional[str] = None
    sector: Optional[str] = None
    fecha_constitucion: Optional[str] = None
    vocero_coordinador: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None

class VoceroCreate(BaseModel):
    consejo_id: Optional[int] = None
    nombre: str
    cedula: Optional[str] = None
    cargo: str
    telefono: Optional[str] = None
    email: Optional[str] = None

# ── Consejos Comunales ────────────────────────────────────────────────────────

@router.get("/consejos")
async def listar_consejos():
    """Lista todos los consejos comunales activos."""
    async with await get_db() as db:
        cursor = await db.execute(
            "SELECT * FROM consejos_comunales WHERE activo=1 ORDER BY nombre"
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]

@router.post("/consejos")
async def crear_consejo(data: ConsejoCreate):
    """Crea un nuevo consejo comunal."""
    async with await get_db() as db:
        cursor = await db.execute(
            """INSERT INTO consejos_comunales 
               (nombre, rif, sector, fecha_constitucion, vocero_coordinador, telefono, email)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (data.nombre, data.rif, data.sector, data.fecha_constitucion,
             data.vocero_coordinador, data.telefono, data.email)
        )
        await db.commit()
        return {"id": cursor.lastrowid, "mensaje": "Consejo comunal creado exitosamente"}

@router.put("/consejos/{consejo_id}")
async def actualizar_consejo(consejo_id: int, data: ConsejoCreate):
    """Actualiza los datos de un consejo comunal."""
    async with await get_db() as db:
        await db.execute(
            """UPDATE consejos_comunales SET nombre=?, rif=?, sector=?, 
               fecha_constitucion=?, vocero_coordinador=?, telefono=?, email=?
               WHERE id=?""",
            (data.nombre, data.rif, data.sector, data.fecha_constitucion,
             data.vocero_coordinador, data.telefono, data.email, consejo_id)
        )
        await db.commit()
        return {"mensaje": "Consejo actualizado exitosamente"}

@router.delete("/consejos/{consejo_id}")
async def eliminar_consejo(consejo_id: int):
    """Desactiva un consejo comunal (baja lógica)."""
    async with await get_db() as db:
        await db.execute(
            "UPDATE consejos_comunales SET activo=0 WHERE id=?", (consejo_id,)
        )
        await db.commit()
        return {"mensaje": "Consejo desactivado exitosamente"}

# ── Voceros ───────────────────────────────────────────────────────────────────

@router.get("/voceros")
async def listar_voceros(consejo_id: Optional[int] = None, q: Optional[str] = None):
    """Lista voceros. Si se pasa consejo_id, filtra por consejo. Si se pasa q, busca por nombre."""
    async with await get_db() as db:
        if q:
            cursor = await db.execute(
                """SELECT v.*, c.nombre as consejo_nombre 
                   FROM voceros v LEFT JOIN consejos_comunales c ON v.consejo_id = c.id
                   WHERE v.activo=1 AND v.nombre LIKE ? ORDER BY v.nombre LIMIT 10""",
                (f"%{q}%",)
            )
        elif consejo_id:
            cursor = await db.execute(
                """SELECT v.*, c.nombre as consejo_nombre 
                   FROM voceros v LEFT JOIN consejos_comunales c ON v.consejo_id = c.id
                   WHERE v.activo=1 AND v.consejo_id=? ORDER BY v.nombre""",
                (consejo_id,)
            )
        else:
            cursor = await db.execute(
                """SELECT v.*, c.nombre as consejo_nombre 
                   FROM voceros v LEFT JOIN consejos_comunales c ON v.consejo_id = c.id
                   WHERE v.activo=1 ORDER BY v.nombre"""
            )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]

@router.post("/voceros")
async def crear_vocero(data: VoceroCreate):
    """Registra un nuevo vocero en el padrón."""
    async with await get_db() as db:
        try:
            cursor = await db.execute(
                """INSERT INTO voceros (consejo_id, nombre, cedula, cargo, telefono, email)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (data.consejo_id, data.nombre, data.cedula, data.cargo,
                 data.telefono, data.email)
            )
            await db.commit()
            return {"id": cursor.lastrowid, "mensaje": "Vocero registrado exitosamente"}
        except Exception as e:
            if "UNIQUE" in str(e):
                raise HTTPException(status_code=400, detail="Ya existe un vocero con esa cédula")
            raise HTTPException(status_code=500, detail=str(e))

@router.put("/voceros/{vocero_id}")
async def actualizar_vocero(vocero_id: int, data: VoceroCreate):
    """Actualiza datos de un vocero."""
    async with await get_db() as db:
        await db.execute(
            """UPDATE voceros SET consejo_id=?, nombre=?, cedula=?, cargo=?, telefono=?, email=?
               WHERE id=?""",
            (data.consejo_id, data.nombre, data.cedula, data.cargo,
             data.telefono, data.email, vocero_id)
        )
        await db.commit()
        return {"mensaje": "Vocero actualizado exitosamente"}

@router.delete("/voceros/{vocero_id}")
async def eliminar_vocero(vocero_id: int):
    """Desactiva un vocero (baja lógica)."""
    async with await get_db() as db:
        await db.execute("UPDATE voceros SET activo=0 WHERE id=?", (vocero_id,))
        await db.commit()
        return {"mensaje": "Vocero desactivado exitosamente"}

# ── Estadísticas ──────────────────────────────────────────────────────────────

@router.get("/estadisticas")
async def estadisticas_padron():
    """Retorna estadísticas resumidas del padrón."""
    async with await get_db() as db:
        c1 = await db.execute("SELECT COUNT(*) FROM consejos_comunales WHERE activo=1")
        c2 = await db.execute("SELECT COUNT(*) FROM voceros WHERE activo=1")
        total_consejos = (await c1.fetchone())[0]
        total_voceros = (await c2.fetchone())[0]
        return {
            "total_consejos": total_consejos,
            "total_voceros": total_voceros
        }

# ── Exportación a Excel ───────────────────────────────────────────────────────

@router.get("/exportar-xlsx")
async def exportar_padron_xlsx():
    """Exporta el padrón completo (consejos y voceros) a un archivo Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ─── Hoja 1: Consejos Comunales ───
        ws1 = wb.active
        ws1.title = "Consejos Comunales"
        
        headers_consejos = ["ID", "Nombre", "RIF", "Sector", "Fecha Constitución",
                            "Vocero Coordinador", "Teléfono", "Email"]
        header_fill = PatternFill("solid", fgColor="1a365d")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, h in enumerate(headers_consejos, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws1.column_dimensions[get_column_letter(col)].width = 20

        async with await get_db() as db:
            cursor = await db.execute(
                "SELECT id, nombre, rif, sector, fecha_constitucion, vocero_coordinador, telefono, email FROM consejos_comunales WHERE activo=1 ORDER BY nombre"
            )
            rows = await cursor.fetchall()
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    ws1.cell(row=r, column=c, value=val)

        # ─── Hoja 2: Voceros ───
        ws2 = wb.create_sheet("Voceros")
        
        headers_voceros = ["ID", "Consejo", "Nombre", "Cédula", "Cargo", "Teléfono", "Email"]
        for col, h in enumerate(headers_voceros, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col)].width = 22

        async with await get_db() as db:
            cursor = await db.execute(
                """SELECT v.id, c.nombre, v.nombre, v.cedula, v.cargo, v.telefono, v.email
                   FROM voceros v LEFT JOIN consejos_comunales c ON v.consejo_id = c.id
                   WHERE v.activo=1 ORDER BY v.nombre"""
            )
            rows = await cursor.fetchall()
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    ws2.cell(row=r, column=c, value=val)

        # ─── Exportar a stream ───
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        filename = f"padron_cppo_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error al exportar padrón a Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
